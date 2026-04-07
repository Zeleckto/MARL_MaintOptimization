"""
analytics/excel_writer.py
=========================
Writes conference-quality Excel workbooks.
Dark background, colour-coded cells, proper formatting.
"""
import os
import numpy as np

try:
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  numbers as xlnums)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

BG       = "FF0A0C12"
HEADER   = "FF1A1D25"
BEST_BG  = "FF0D3B1F"
BEST_FG  = "FF00DC6E"
WORST_BG = "FF3B0D0D"
WORST_FG = "FFFF4444"
MARL_BG  = "FF001E2B"
MARL_FG  = "FF00C8FF"
TEXT_FG  = "FFFAFAFA"
SUB_FG   = "FF8899AA"
GOLD     = "FFFFB900"

BORDER = Border(
    left=Side(style="thin", color="FF2A2D35"),
    right=Side(style="thin", color="FF2A2D35"),
    top=Side(style="thin", color="FF2A2D35"),
    bottom=Side(style="thin", color="FF2A2D35"),
)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(hex_color=TEXT_FG, bold=False, size=10):
    return Font(color=hex_color, bold=bold, size=size, name="Calibri")

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _cell(ws, row, col, value, bg=BG, fg=TEXT_FG, bold=False, fmt=None, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = _fill(bg)
    c.font      = _font(fg, bold=bold, size=size)
    c.alignment = _align()
    c.border    = BORDER
    if fmt:
        c.number_format = fmt
    return c


def write_comparison_workbook(
    results: dict,          # {policy_name: [kpi_dict, ...]} episode lists
    outpath: str,
    title: str = "MARL vs Baselines",
    lower_better: set = None,
    higher_better: set = None,
):
    """
    Write a full comparison workbook with sheets:
      1. Summary — mean±std for all policies, colour-coded best/worst
      2. Episodes — raw per-episode data
      3. Statistical — Wilcoxon + Cohen's d vs ABR (best baseline)
      4. Reliability — reliability engineering KPIs table
      5. Scheduling — JSS KPIs table
      6. Cost — cost decomposition table
    """
    if not EXCEL_OK:
        print("openpyxl not available — skipping Excel output"); return

    if lower_better is None:
        lower_better = {"failures","weighted_tardiness","wt_normalised","makespan",
                        "mean_flow_time","n_CM","failure_cost","total_cost",
                        "holding_cost","order_cost","maint_cost","mean_hazard_rate",
                        "health_cv","mttr"}
    if higher_better is None:
        higher_better = {"jobs_completed","service_level","availability",
                         "inherent_availability","n_PM","pm_cm_ratio","mtbf",
                         "mean_health","mean_rul_norm","throughput",
                         "machine_utilisation","fill_rate"}

    from scipy import stats as sci_stats

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    policies = list(results.keys())
    marl_key = next((k for k in policies if "MARL" in k or "marl" in k.lower()), None)
    abr_key  = next((k for k in policies if "ABR"  in k), None)
    all_kpis = list(results[policies[0]][0].keys())

    # ── SHEET 1: Summary ───────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    for col in range(1, 3 + len(policies) * 2 + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    # Title row
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=2 + len(policies) * 2)
    tc = ws.cell(1, 1, title)
    tc.fill = _fill("FF101420"); tc.font = _font(MARL_FG, bold=True, size=14)
    tc.alignment = _align(h="left")

    # Header: Domain | KPI | Policy1 mean±std | Policy2 ...
    row = 3
    _cell(ws, row, 1, "Domain", HEADER, GOLD, bold=True)
    _cell(ws, row, 2, "KPI", HEADER, GOLD, bold=True)
    for i, p in enumerate(policies):
        bg = MARL_BG if p == marl_key else HEADER
        fg = MARL_FG if p == marl_key else GOLD
        _cell(ws, row, 3 + i*2, f"{p} mean", bg, fg, bold=True)
        _cell(ws, row, 4 + i*2, f"± std",   bg, fg, bold=True)

    # KPI groups for paper
    kpi_groups = [
        ("Scheduling", [
            ("jobs_completed",     "Jobs Completed",       "{:.1f}"),
            ("service_level",      "Service Level",        "{:.3f}"),
            ("weighted_tardiness", "Wt. Tardiness",        "{:.1f}"),
            ("makespan",           "Makespan (shifts)",    "{:.1f}"),
            ("mean_flow_time",     "Mean Flow Time",       "{:.1f}"),
            ("machine_utilisation","Machine Utilisation",  "{:.3f}"),
        ]),
        ("Reliability", [
            ("availability",          "Availability",          "{:.4f}"),
            ("inherent_availability", "Inherent Avail. (ISO)", "{:.4f}"),
            ("failures",              "Failures / ep",         "{:.2f}"),
            ("n_PM",                  "PM Events / ep",        "{:.2f}"),
            ("n_CM",                  "CM Events / ep",        "{:.2f}"),
            ("pm_cm_ratio",           "PM/CM Ratio",           "{:.2f}"),
            ("mtbf",                  "MTBF (shifts)",         "{:.1f}"),
            ("mean_health",           "Mean Health (%)",       "{:.1f}"),
            ("mean_rul_norm",         "Mean RUL (norm)",       "{:.3f}"),
        ]),
        ("Cost & Inventory", [
            ("total_cost",    "Total Cost",        "{:.1f}"),
            ("maint_cost",    "Maintenance Cost",  "{:.1f}"),
            ("failure_cost",  "Failure Cost",      "{:.1f}"),
            ("order_cost",    "Order Cost",        "{:.1f}"),
            ("holding_cost",  "Holding Cost",      "{:.2f}"),
            ("fill_rate",     "Inventory Fill Rate","{:.3f}"),
            ("eoq_ratio",     "EOQ Ratio",         "{:.2f}"),
        ]),
    ]

    row = 4
    # Compute means/stds
    means = {p: {k: float(np.mean([e[k] for e in eps if k in e]))
                 for k in all_kpis}
             for p, eps in results.items()}
    stds  = {p: {k: float(np.std([e[k]  for e in eps if k in e], ddof=1))
                 for k in all_kpis}
             for p, eps in results.items()}

    for domain, kpis in kpi_groups:
        first = True
        for (key, label, fmt) in kpis:
            if key not in all_kpis:
                continue
            bg_d = "FF0E1018" if first else BG
            _cell(ws, row, 1, domain if first else "", bg_d, SUB_FG)
            first = False
            _cell(ws, row, 2, label, BG, TEXT_FG)

            # Find best/worst value
            vals = [(p, means[p][key]) for p in policies if key in means[p]]
            if not vals:
                row += 1; continue
            best_p = min(vals, key=lambda x: x[1])[0] if key in lower_better \
                else max(vals, key=lambda x: x[1])[0]
            worst_p = max(vals, key=lambda x: x[1])[0] if key in lower_better \
                else min(vals, key=lambda x: x[1])[0]

            for i, p in enumerate(policies):
                m = means[p].get(key, 0)
                s = stds[p].get(key, 0)
                if p == best_p:
                    bg_c, fg_c = BEST_BG, BEST_FG
                elif p == worst_p:
                    bg_c, fg_c = WORST_BG, WORST_FG
                elif p == marl_key:
                    bg_c, fg_c = MARL_BG, MARL_FG
                else:
                    bg_c, fg_c = BG, TEXT_FG
                _cell(ws, row, 3 + i*2, float(f"{m:.4f}"), bg_c, fg_c)
                _cell(ws, row, 4 + i*2, float(f"{s:.4f}"), bg_c, SUB_FG)
            row += 1

    # ── SHEET 2: Raw Episodes ──────────────────────────────────────
    ws2 = wb.create_sheet("Episodes")
    ws2.sheet_view.showGridLines = False
    headers = ["policy", "episode"] + all_kpis
    for ci, h in enumerate(headers, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = 16
        _cell(ws2, 1, ci, h, HEADER, GOLD, bold=True)
    row = 2
    for p, eps in results.items():
        for ep_i, ep in enumerate(eps):
            _cell(ws2, row, 1, p, BG, TEXT_FG)
            _cell(ws2, row, 2, ep_i + 1, BG, TEXT_FG)
            for ci, k in enumerate(all_kpis, 3):
                _cell(ws2, row, ci, float(ep.get(k, 0)), BG, TEXT_FG, fmt="#,##0.0000")
            row += 1

    # ── SHEET 3: Statistical (Wilcoxon vs ABR) ────────────────────
    ws3 = wb.create_sheet("Statistics")
    ws3.sheet_view.showGridLines = False
    stat_header = ["KPI", "MARL mean", "ABR mean", "Δ%",
                   "Wilcoxon p", "Significant?", "Cohen's d", "Effect size"]
    for ci, h in enumerate(stat_header, 1):
        ws3.column_dimensions[get_column_letter(ci)].width = 18
        _cell(ws3, 1, ci, h, HEADER, GOLD, bold=True)

    if marl_key and abr_key:
        row3 = 2
        for key in all_kpis:
            a = [e[key] for e in results[marl_key] if key in e]
            b = [e[key] for e in results[abr_key]  if key in e]
            if not a or not b: continue
            ma, mb = np.mean(a), np.mean(b)
            delta_pct = (ma - mb) / max(abs(mb), 1e-8) * 100
            try:
                stat, pval = sci_stats.wilcoxon(a[:len(b)], b[:len(a)])
            except:
                pval = 1.0
            # Cohen's d
            na, nb = len(a), len(b)
            pool_std = np.sqrt(((na-1)*np.std(a,ddof=1)**2 + (nb-1)*np.std(b,ddof=1)**2) / (na+nb-2))
            d = abs(ma - mb) / max(pool_std, 1e-10)
            d_interp = "large" if d > 0.8 else ("medium" if d > 0.5 else ("small" if d > 0.2 else "negligible"))
            sig = "YES ***" if pval < 0.001 else ("YES **" if pval < 0.01 else ("YES *" if pval < 0.05 else "no"))
            sig_fg = BEST_FG if pval < 0.05 else WORST_FG
            _cell(ws3, row3, 1, key, BG, TEXT_FG)
            _cell(ws3, row3, 2, round(ma, 4), MARL_BG, MARL_FG)
            _cell(ws3, row3, 3, round(mb, 4), BG, TEXT_FG)
            _cell(ws3, row3, 4, round(delta_pct, 1), BG, (BEST_FG if delta_pct > 0 else WORST_FG))
            _cell(ws3, row3, 5, round(pval, 5), BG, TEXT_FG)
            _cell(ws3, row3, 6, sig, BG, sig_fg, bold=(pval < 0.05))
            _cell(ws3, row3, 7, round(d, 3), BG, TEXT_FG)
            _cell(ws3, row3, 8, d_interp, BG, GOLD)
            row3 += 1

    os.makedirs(os.path.dirname(outpath) if os.path.dirname(outpath) else ".", exist_ok=True)
    wb.save(outpath)
    print(f"  Saved: {outpath}")
