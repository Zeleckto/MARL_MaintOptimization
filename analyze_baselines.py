"""
analyze_baselines.py
=====================
Runs all 4 baselines across multiple episodes and generates:
  1. Comparison bar chart (PNG) — all metrics side by side
  2. Radar / spider chart (PNG) — multi-metric profile per baseline
  3. Excel report (2 sheets):
       Sheet 1: Raw per-episode data for all baselines
       Sheet 2: Mean ± Std comparison table

Optionally adds trained MARL policy results if a checkpoint is provided.

Usage:
    python analyze_baselines.py                               # baselines only
    python analyze_baselines.py --episodes 20                 # more episodes
    python analyze_baselines.py --checkpoint checkpoints/latest.pt  # + MARL

Output:
    results/baseline_comparison.png
    results/baseline_radar.png
    results/baseline_comparison.xlsx
"""

import argparse
import os
import sys
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from typing import List, Dict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import yaml
from environments.mfg_env import ManufacturingEnv, AGENT_PDM, AGENT_JOBSHOP
from benchmarks.baselines import get_all_baselines, BaselinePolicy

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


# ---------------------------------------------------------------------------
# COLOUR PALETTE — one colour per baseline + MARL
# ---------------------------------------------------------------------------
BASELINE_COLOURS = [
    "#FF3246",   # Reactive — red (bad baseline)
    "#FFB900",   # Rule-based EDF — amber
    "#B43CFF",   # Fixed-interval — purple
    "#00DC6E",   # ABR+MDD+(Q,R) — green (analytical best)
    "#00C8FF",   # MARL — cyan (our method)
]

BG    = "#0A0C12"
GRID  = "#1A1E28"
TEXT  = "#D2DCF7"


# ---------------------------------------------------------------------------
# RUN BASELINES
# ---------------------------------------------------------------------------
METRICS = [
    ("failures",      "Failures/Ep",    "lower",  True),
    ("n_PM",          "PM Events",      "higher", False),
    ("n_CM",          "CM Events",      "lower",  True),
    ("pm_cm_ratio",   "PM/CM Ratio",    "higher", False),
    ("completions",   "Jobs Completed", "higher", False),
    ("tardiness",     "Wt. Tardiness",  "lower",  True),
    ("service_level", "Service Level",  "higher", False),
    ("avg_health",    "Avg Health %",   "higher", False),
]


def run_one_episode(env, baseline, seed):
    baseline.reset()
    env.reset(seed=seed)
    done   = False
    steps  = 0
    n_PM   = 0
    n_CM   = 0

    while not done and steps < 300:
        a1 = baseline.agent1_action(env)
        env._step_agent1(a1)
        a2 = baseline.agent2_action(env)
        env._step_agent2(a2)
        env._resolve_physics()
        env._compute_rewards()

        n_PM += sum(1 for a in a1["maintenance"] if a == 1)
        n_CM += sum(1 for a in a1["maintenance"] if a == 2)

        done = env.terminations[AGENT_PDM] or env.truncations[AGENT_PDM]
        steps += 1

    completed = [j for j in env.jobs if j.completion_time is not None]
    on_time   = [j for j in completed if j.tardiness == 0]
    tard      = sum(j.weight * j.tardiness for j in completed)
    avg_health = float(np.mean([s.health for s in env.machine_states]))

    return {
        "failures":      env._episode_failures,
        "n_PM":          n_PM,
        "n_CM":          n_CM,
        "pm_cm_ratio":   n_PM / max(n_CM, 1),
        "completions":   len(completed),
        "tardiness":     tard,
        "service_level": len(on_time) / max(len(completed), 1),
        "avg_health":    avg_health,
    }


def collect_results(env, baselines, n_episodes, seed_offset):
    """Returns {baseline_name: [episode_metrics_dict]}"""
    all_results = {}
    for b in baselines:
        print(f"  Running: {b.name}")
        episodes = []
        for ep in range(n_episodes):
            m = run_one_episode(env, b, seed=seed_offset + ep * 7)
            episodes.append(m)
            sys.stdout.write(f"\r    Episode {ep+1}/{n_episodes}")
            sys.stdout.flush()
        print()
        all_results[b.name] = episodes
    return all_results


def summarise(episodes: List[Dict]) -> Dict:
    keys = list(episodes[0].keys())
    return {k: (float(np.mean([e[k] for e in episodes])),
                float(np.std([e[k] for e in episodes])))
            for k in keys}


# ---------------------------------------------------------------------------
# PLOTS
# ---------------------------------------------------------------------------
def make_comparison_bar_chart(summaries: dict, output_path: str) -> None:
    """Bar chart comparing all baselines across all metrics."""
    baseline_names = list(summaries.keys())
    n_b = len(baseline_names)
    n_m = len(METRICS)

    fig, axes = plt.subplots(2, 4, figsize=(20, 9))
    fig.patch.set_facecolor(BG)
    axes = axes.flatten()

    for mi, (key, label, direction, _) in enumerate(METRICS):
        ax = axes[mi]
        ax.set_facecolor(GRID)
        ax.tick_params(colors=TEXT, labelsize=8)
        ax.title.set_color(TEXT)
        for spine in ax.spines.values():
            spine.set_edgecolor(GRID)
        ax.grid(True, axis="y", color="#2A2E3A", linewidth=0.5)

        means  = [summaries[b][key][0] for b in baseline_names]
        stds   = [summaries[b][key][1] for b in baseline_names]
        colours = BASELINE_COLOURS[:n_b]

        bars = ax.bar(range(n_b), means, color=colours, alpha=0.85,
                      yerr=stds, capsize=4, error_kw={"color": TEXT, "linewidth": 1})

        # Star best bar
        best = min(means) if direction == "lower" else max(means)
        for i, (mean, bar) in enumerate(zip(means, bars)):
            if abs(mean - best) < 1e-6:
                ax.text(i, mean + stds[i] * 0.1 + best * 0.02,
                        "★", ha="center", va="bottom", fontsize=12,
                        color=colours[i])

        ax.set_title(label, fontsize=9, fontweight="bold")
        ax.set_xticks([])
        ax.set_xlabel("")

    # Legend
    handles = [plt.Rectangle((0,0),1,1, color=BASELINE_COLOURS[i], alpha=0.85)
               for i in range(n_b)]
    short_names = [n.replace(" + ", "+").replace("Fixed-Interval PM + SPT", "Fixed-Int+SPT")
                   for n in baseline_names]
    fig.legend(handles, short_names,
               loc="lower center", ncol=n_b,
               facecolor=GRID, labelcolor=TEXT, fontsize=9,
               bbox_to_anchor=(0.5, 0.01))

    fig.suptitle("BTP2 — Baseline Comparison\n(★ = best for each metric)",
                 fontsize=13, fontweight="bold", color=TEXT, y=0.99)

    plt.tight_layout(rect=[0, 0.07, 1, 0.97])
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    plt.savefig(output_path, dpi=150, bbox_inches="tight", facecolor=BG)
    plt.close()
    print(f"Saved: {output_path}")


def make_radar_chart(summaries: dict, output_path: str) -> None:
    """Radar/spider chart — multi-metric profile per baseline."""
    # Use metrics that can be normalised to [0,1] where higher=better
    radar_metrics = [
        ("service_level", "Service\nLevel",    "higher"),
        ("avg_health",    "Avg\nHealth",        "higher"),
        ("pm_cm_ratio",   "PM/CM\nRatio",       "higher"),
        ("completions",   "Jobs\nCompleted",    "higher"),
        ("failures",      "Failures\n(inv)",    "lower"),   # inverted
        ("tardiness",     "Tardiness\n(inv)",   "lower"),   # inverted
    ]

    baseline_names = list(summaries.keys())
    n = len(radar_metrics)
    angles = np.linspace(0, 2 * np.pi, n, endpoint=False).tolist()
    angles += angles[:1]  # close the polygon

    fig, ax = plt.subplots(1, 1, figsize=(8, 8),
                           subplot_kw=dict(polar=True))
    fig.patch.set_facecolor(BG)
    ax.set_facecolor(GRID)

    # Normalise each metric to [0,1] across baselines
    def normalise(key, direction):
        vals = [summaries[b][key][0] for b in baseline_names]
        mn, mx = min(vals), max(vals)
        if mx == mn:
            return [0.5] * len(vals)
        normed = [(v - mn) / (mx - mn) for v in vals]
        if direction == "lower":
            normed = [1.0 - v for v in normed]
        return normed

    normalised = {}
    for key, _, direction in radar_metrics:
        normalised[key] = normalise(key, direction)

    for bi, bname in enumerate(baseline_names):
        values = [normalised[key][bi] for key, _, _ in radar_metrics]
        values += values[:1]
        col = BASELINE_COLOURS[bi]
        ax.plot(angles, values, color=col, linewidth=2, alpha=0.9)
        ax.fill(angles, values, color=col, alpha=0.12)

    ax.set_xticks(angles[:-1])
    ax.set_xticklabels([label for _, label, _ in radar_metrics],
                       color=TEXT, size=9)
    ax.set_yticks([0.25, 0.5, 0.75, 1.0])
    ax.set_yticklabels(["0.25", "0.5", "0.75", "1.0"], color=TEXT, size=7)
    ax.tick_params(colors=TEXT)
    ax.grid(color="#2A2E3A", linewidth=0.8)
    ax.spines["polar"].set_color("#2A2E3A")

    # Legend
    handles = [plt.Line2D([0],[0], color=BASELINE_COLOURS[i], linewidth=2)
               for i in range(len(baseline_names))]
    short = [n.replace(" + ", "+").replace("Fixed-Interval PM + SPT", "Fixed+SPT")
             for n in baseline_names]
    ax.legend(handles, short, loc="upper right",
              bbox_to_anchor=(1.35, 1.1),
              facecolor=GRID, labelcolor=TEXT, fontsize=8)

    ax.set_title("Multi-Metric Profile\n(normalised, higher = better)",
                 color=TEXT, size=11, fontweight="bold", pad=20)

    plt.tight_layout()
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    plt.savefig(output_path, dpi=150, bbox_inches="tight", facecolor=BG)
    plt.close()
    print(f"Saved: {output_path}")


# ---------------------------------------------------------------------------
# EXCEL EXPORT
# ---------------------------------------------------------------------------
def make_baseline_excel(all_results: dict, summaries: dict, output_path: str) -> None:
    if not EXCEL_AVAILABLE:
        print("openpyxl not installed — skipping Excel export")
        return

    wb = openpyxl.Workbook()

    hdr_fill  = PatternFill("solid", fgColor="1A1E28")
    hdr_font  = Font(bold=True, color="00DC6E", size=10)
    data_font = Font(color="D2DCF7", size=9)
    good_fill = PatternFill("solid", fgColor="0D2B1A")
    bad_fill  = PatternFill("solid", fgColor="2B0D0D")
    center    = Alignment(horizontal="center")
    thin = Border(
        left=Side(style="thin", color="2A2E3A"),
        right=Side(style="thin", color="2A2E3A"),
        top=Side(style="thin", color="2A2E3A"),
        bottom=Side(style="thin", color="2A2E3A"),
    )

    def hdr(cell):
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = center; cell.border = thin

    def dat(cell, good=None):
        cell.font = data_font; cell.alignment = center; cell.border = thin
        if good is True:  cell.fill = good_fill
        elif good is False: cell.fill = bad_fill

    # ---- SHEET 1: Raw Episode Data ----
    ws1 = wb.active
    ws1.title = "Raw Episode Data"
    ws1.sheet_properties.tabColor = "00DC6E"

    baseline_names = list(all_results.keys())
    metric_keys    = [k for k, _, _, _ in METRICS]
    metric_labels  = [l for _, l, _, _ in METRICS]

    # Header: Baseline | Episode | Metric1 | Metric2 | ...
    headers = ["Baseline", "Episode"] + metric_labels
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        hdr(cell)
        ws1.column_dimensions[get_column_letter(col)].width = 18

    row_i = 2
    for bname, episodes in all_results.items():
        for ep_i, ep in enumerate(episodes):
            row_data = [bname, ep_i + 1] + [ep.get(k, "") for k in metric_keys]
            for col, val in enumerate(row_data, 1):
                cell = ws1.cell(row=row_i, column=col, value=val)
                dat(cell)
            row_i += 1

    ws1.freeze_panes = "A2"

    # ---- SHEET 2: Mean ± Std Comparison ----
    ws2 = wb.create_sheet("Comparison Table")
    ws2.sheet_properties.tabColor = "00C8FF"

    ws2.cell(row=1, column=1, value="BTP2 — Baseline Comparison (Mean ± Std)")
    ws2["A1"].font = Font(bold=True, color="00DC6E", size=12)
    ws2.merge_cells(f"A1:{get_column_letter(len(baseline_names) * 2 + 1)}1")

    # Header row
    ws2.cell(row=3, column=1, value="Metric").font = hdr_font
    ws2["A3"].fill = hdr_fill; ws2["A3"].border = thin
    ws2.column_dimensions["A"].width = 20

    for bi, bname in enumerate(baseline_names):
        short = bname.replace(" + ", "+").replace("Fixed-Interval PM + SPT", "Fixed+SPT")
        col_mean = 2 + bi * 2
        col_std  = 3 + bi * 2
        cell_m = ws2.cell(row=3, column=col_mean, value=f"{short}\nMean")
        cell_s = ws2.cell(row=3, column=col_std,  value=f"{short}\nStd")
        for c in (cell_m, cell_s):
            hdr(c)
            ws2.column_dimensions[get_column_letter(c.column)].width = 14

    # Data rows
    for mi, (key, label, direction, lower_better) in enumerate(METRICS):
        row = 4 + mi
        ws2.cell(row=row, column=1, value=label)
        ws2["A" + str(row)].font = data_font
        ws2["A" + str(row)].border = thin

        means = [summaries[b][key][0] for b in baseline_names]
        best  = min(means) if direction == "lower" else max(means)

        for bi, bname in enumerate(baseline_names):
            mean, std = summaries[bname][key]
            is_best = abs(mean - best) < 1e-6

            col_mean = 2 + bi * 2
            col_std  = 3 + bi * 2
            cm = ws2.cell(row=row, column=col_mean, value=round(mean, 3))
            cs = ws2.cell(row=row, column=col_std,  value=f"±{std:.3f}")
            dat(cm, good=is_best)
            dat(cs, good=is_best)
            if is_best:
                cm.value = f"★ {round(mean, 3)}"

    ws2.freeze_panes = "B4"

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    print(f"Saved: {output_path}")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--episodes",   type=int, default=10)
    parser.add_argument("--seed",       type=int, default=42)
    parser.add_argument("--jobs",       type=int, default=None)
    parser.add_argument("--outdir",     default="results/")
    parser.add_argument("--no-excel",   action="store_true")
    parser.add_argument("--no-plots",   action="store_true")
    args = parser.parse_args()

    print("\nLoading config...")
    with open("configs/base.yaml") as f:
        config = yaml.safe_load(f)

    if args.jobs:
        config["jobs"]["n_jobs_train"] = args.jobs

    env       = ManufacturingEnv(config)
    baselines = get_all_baselines()

    print(f"\nRunning {len(baselines)} baselines × {args.episodes} episodes each...")
    all_results = collect_results(env, baselines, args.episodes, args.seed)
    summaries   = {b: summarise(eps) for b, eps in all_results.items()}

    print("\n" + "="*60)
    print("  QUICK RESULTS TABLE")
    print("="*60)
    name_w = 25
    print(f"  {'Metric':<20} " + "  ".join(f"{b[:name_w]:<{name_w}}" for b in summaries))
    print("  " + "-"*90)
    for key, label, direction, _ in METRICS:
        means = [summaries[b][key][0] for b in summaries]
        best  = min(means) if direction == "lower" else max(means)
        row = f"  {label:<20} "
        for b, mean in zip(summaries, means):
            marker = "★" if abs(mean - best) < 1e-6 else " "
            row += f"  {marker}{mean:>8.2f}          "
        print(row)
    print()

    os.makedirs(args.outdir, exist_ok=True)

    if not args.no_plots:
        print("Generating comparison bar chart...")
        make_comparison_bar_chart(summaries, os.path.join(args.outdir, "baseline_comparison.png"))
        print("Generating radar chart...")
        make_radar_chart(summaries, os.path.join(args.outdir, "baseline_radar.png"))

    if not args.no_excel:
        print("Generating Excel report...")
        make_baseline_excel(all_results, summaries, os.path.join(args.outdir, "baseline_comparison.xlsx"))

    print(f"\nAll outputs in: {args.outdir}/")
    print("  baseline_comparison.png  — bar chart, all metrics")
    print("  baseline_radar.png       — spider chart, multi-metric profile")
    print("  baseline_comparison.xlsx — raw data + comparison table")


if __name__ == "__main__":
    main()
