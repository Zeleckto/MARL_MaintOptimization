"""
analyze_training.py
====================
Reads TensorBoard event files and generates:
  1. Training curve plots (PNG) — rewards, losses, entropy, KPIs
  2. Excel report (3 sheets):
       Sheet 1: Raw per-episode data
       Sheet 2: Early vs Late stage comparison (first 20% vs last 20% of training)
       Sheet 3: Summary statistics

Usage:
    python analyze_training.py                         # reads runs/ folder
    python analyze_training.py --logdir runs/phase1_*  # specific run
    python analyze_training.py --logdir runs/phase1_1745000000

Output:
    results/training_curves.png
    results/training_analysis.xlsx
"""

import argparse
import os
import glob
import numpy as np
import matplotlib
matplotlib.use("Agg")  # headless
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from collections import defaultdict

try:
    from tensorboard.backend.event_processing.event_accumulator import EventAccumulator
    TB_AVAILABLE = True
except ImportError:
    TB_AVAILABLE = False
    print("WARNING: tensorboard not installed — install with: pip install tensorboard")

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("WARNING: openpyxl not installed — Excel export disabled")


# ---------------------------------------------------------------------------
# COLOUR PALETTE (matches viz_factory.py aesthetic)
# ---------------------------------------------------------------------------
COLOURS = {
    "r1":        "#00DC6E",   # neon green  — Agent 1
    "r2":        "#00C8FF",   # cyan        — Agent 2
    "shared":    "#FF1E37",   # red         — failures
    "entropy":   "#FFB900",   # amber
    "loss":      "#B43CFF",   # purple
    "health":    "#00DC6E",
    "tard":      "#FF3246",
    "failures":  "#FF1E37",
    "bg":        "#0A0C12",
    "grid":      "#1A1E28",
    "text":      "#D2DCF7",
}

TAGS = {
    "rewards": [
        ("rewards/r1",     "Agent 1 Return (r1)",   COLOURS["r1"]),
        ("rewards/r2",     "Agent 2 Return (r2)",   COLOURS["r2"]),
        ("rewards/shared", "Shared (failures)",      COLOURS["shared"]),
    ],
    "episode": [
        ("episode/return1",           "Ep Return Agent 1",  COLOURS["r1"]),
        ("episode/return2",           "Ep Return Agent 2",  COLOURS["r2"]),
        ("episode/failures",          "Failures/Episode",   COLOURS["failures"]),
        ("episode/weighted_tardiness","Weighted Tardiness",  COLOURS["tard"]),
        ("episode/jobs_completed",    "Jobs Completed",      COLOURS["health"]),
        ("episode/avg_health",        "Avg Machine Health",  COLOURS["health"]),
        ("episode/service_level",     "Service Level",       COLOURS["r2"]),
        ("episode/pm_cm_ratio",       "PM/CM Ratio",         COLOURS["entropy"]),
    ],
    "training": [
        ("train/actor1_loss", "Actor 1 Loss",  COLOURS["r1"]),
        ("train/actor2_loss", "Actor 2 Loss",  COLOURS["r2"]),
        ("train/critic_loss", "Critic Loss",   COLOURS["loss"]),
        ("train/entropy1",    "Entropy 1",     COLOURS["entropy"]),
        ("train/entropy2",    "Entropy 2",     COLOURS["entropy"]),
    ],
}


# ---------------------------------------------------------------------------
# DATA LOADING
# ---------------------------------------------------------------------------
def load_tb_data(logdir: str) -> dict:
    """
    Loads all scalar data from TensorBoard event files.
    Returns {tag: [(step, value), ...]}
    """
    if not TB_AVAILABLE:
        return {}

    # Find event files
    event_files = glob.glob(os.path.join(logdir, "**", "events.out.tfevents.*"),
                            recursive=True)
    if not event_files:
        event_files = glob.glob(os.path.join(logdir, "events.out.tfevents.*"))

    if not event_files:
        print(f"No TensorBoard event files found in: {logdir}")
        return {}

    print(f"Found {len(event_files)} event file(s)")

    all_data = defaultdict(list)
    for ef in event_files:
        ea = EventAccumulator(ef)
        ea.Reload()
        for tag in ea.Tags().get("scalars", []):
            events = ea.Scalars(tag)
            all_data[tag].extend([(e.step, e.value) for e in events])

    # Sort by step
    for tag in all_data:
        all_data[tag].sort(key=lambda x: x[0])

    print(f"Loaded {len(all_data)} scalar tags:")
    for tag in sorted(all_data.keys()):
        print(f"  {tag}: {len(all_data[tag])} points")

    return dict(all_data)


def smooth(values, window=20):
    """Moving average smoothing."""
    if len(values) < window:
        return values
    kernel = np.ones(window) / window
    return np.convolve(values, kernel, mode="valid")


# ---------------------------------------------------------------------------
# PLOTTING
# ---------------------------------------------------------------------------
def make_training_plots(data: dict, output_path: str) -> None:
    """
    Generates training curve figure with 4 panels:
      Panel 1: Episode returns (r1 and r2 over episodes)
      Panel 2: Failures + Tardiness per episode
      Panel 3: Machine health + Service level
      Panel 4: PPO losses + Entropy
    """
    fig = plt.figure(figsize=(18, 12))
    fig.patch.set_facecolor(COLOURS["bg"])
    gs = gridspec.GridSpec(2, 2, figure=fig, hspace=0.45, wspace=0.35)

    axes = [fig.add_subplot(gs[i, j]) for i in range(2) for j in range(2)]
    for ax in axes:
        ax.set_facecolor(COLOURS["grid"])
        ax.tick_params(colors=COLOURS["text"], labelsize=9)
        ax.xaxis.label.set_color(COLOURS["text"])
        ax.yaxis.label.set_color(COLOURS["text"])
        ax.title.set_color(COLOURS["text"])
        for spine in ax.spines.values():
            spine.set_edgecolor(COLOURS["grid"])
        ax.grid(True, color="#2A2E3A", linewidth=0.5, alpha=0.7)

    # ---- Panel 1: Episode Returns ----
    ax = axes[0]
    ax.set_title("Episode Returns", fontsize=11, fontweight="bold")
    for tag, label, colour in TAGS["episode"][:2]:
        if tag in data:
            steps, vals = zip(*data[tag])
            ax.plot(steps, vals, alpha=0.2, color=colour, linewidth=0.5)
            sm = smooth(list(vals))
            sm_steps = list(steps)[len(steps)-len(sm):]
            ax.plot(sm_steps, sm, color=colour, linewidth=1.8, label=label)
    ax.set_xlabel("Episode")
    ax.set_ylabel("Return")
    ax.legend(fontsize=8, facecolor=COLOURS["grid"], labelcolor=COLOURS["text"])

    # ---- Panel 2: Failures + Tardiness ----
    ax2 = axes[1]
    ax2.set_title("Failures & Tardiness per Episode", fontsize=11, fontweight="bold")
    ax2b = ax2.twinx()
    ax2b.set_facecolor(COLOURS["grid"])
    ax2b.tick_params(colors=COLOURS["text"], labelsize=9)

    if "episode/failures" in data:
        steps, vals = zip(*data["episode/failures"])
        ax2.plot(steps, vals, alpha=0.2, color=COLOURS["failures"], linewidth=0.5)
        sm = smooth(list(vals))
        sm_steps = list(steps)[len(steps)-len(sm):]
        ax2.plot(sm_steps, sm, color=COLOURS["failures"], linewidth=1.8, label="Failures")
    ax2.set_xlabel("Episode")
    ax2.set_ylabel("Failures", color=COLOURS["failures"])

    if "episode/weighted_tardiness" in data:
        steps, vals = zip(*data["episode/weighted_tardiness"])
        ax2b.plot(steps, vals, alpha=0.2, color=COLOURS["tard"], linewidth=0.5)
        sm = smooth(list(vals))
        sm_steps = list(steps)[len(steps)-len(sm):]
        ax2b.plot(sm_steps, sm, color=COLOURS["tard"], linewidth=1.8,
                  linestyle="--", label="Tardiness")
    ax2b.set_ylabel("Tardiness", color=COLOURS["tard"])
    ax2b.yaxis.label.set_color(COLOURS["tard"])
    ax2b.tick_params(colors=COLOURS["tard"])

    # ---- Panel 3: Health + Service Level ----
    ax = axes[2]
    ax.set_title("Machine Health & Service Level", fontsize=11, fontweight="bold")
    ax3b = ax.twinx()
    ax3b.set_facecolor(COLOURS["grid"])
    ax3b.tick_params(colors=COLOURS["text"], labelsize=9)

    if "episode/avg_health" in data:
        steps, vals = zip(*data["episode/avg_health"])
        ax.plot(steps, vals, alpha=0.2, color=COLOURS["health"], linewidth=0.5)
        sm = smooth(list(vals))
        sm_steps = list(steps)[len(steps)-len(sm):]
        ax.plot(sm_steps, sm, color=COLOURS["health"], linewidth=1.8, label="Avg Health %")
    ax.set_xlabel("Episode")
    ax.set_ylabel("Health %", color=COLOURS["health"])
    ax.set_ylim(0, 100)

    if "episode/service_level" in data:
        steps, vals = zip(*data["episode/service_level"])
        ax3b.plot(steps, vals, alpha=0.2, color=COLOURS["r2"], linewidth=0.5)
        sm = smooth(list(vals))
        sm_steps = list(steps)[len(steps)-len(sm):]
        ax3b.plot(sm_steps, sm, color=COLOURS["r2"], linewidth=1.8,
                  linestyle="--", label="Service Level")
    ax3b.set_ylabel("Service Level", color=COLOURS["r2"])
    ax3b.yaxis.label.set_color(COLOURS["r2"])
    ax3b.tick_params(colors=COLOURS["r2"])
    ax3b.set_ylim(0, 1)

    # ---- Panel 4: Losses + Entropy ----
    ax = axes[3]
    ax.set_title("PPO Losses & Entropy", fontsize=11, fontweight="bold")
    ax4b = ax.twinx()
    ax4b.set_facecolor(COLOURS["grid"])
    ax4b.tick_params(colors=COLOURS["text"], labelsize=9)

    for tag, label, colour in TAGS["training"][:3]:
        if tag in data:
            steps, vals = zip(*data[tag])
            sm = smooth(list(vals))
            sm_steps = list(steps)[len(steps)-len(sm):]
            ax.plot(sm_steps, sm, linewidth=1.5, label=label, color=colour)
    ax.set_xlabel("Step")
    ax.set_ylabel("Loss")
    ax.legend(fontsize=7, facecolor=COLOURS["grid"], labelcolor=COLOURS["text"])

    for tag, label, colour in TAGS["training"][3:]:
        if tag in data:
            steps, vals = zip(*data[tag])
            sm = smooth(list(vals))
            sm_steps = list(steps)[len(steps)-len(sm):]
            ax4b.plot(sm_steps, sm, linewidth=1.5, linestyle="--",
                      label=label, color=COLOURS["entropy"])
    ax4b.set_ylabel("Entropy", color=COLOURS["entropy"])
    ax4b.yaxis.label.set_color(COLOURS["entropy"])
    ax4b.tick_params(colors=COLOURS["entropy"])

    # Title
    fig.suptitle(
        "BTP2 — MARL Manufacturing Training Analysis",
        fontsize=14, fontweight="bold", color=COLOURS["text"], y=0.98
    )

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    plt.savefig(output_path, dpi=150, bbox_inches="tight", facecolor=COLOURS["bg"])
    plt.close()
    print(f"Saved: {output_path}")


def make_early_vs_late_plot(data: dict, output_path: str) -> None:
    """
    Side-by-side comparison: first 20% vs last 20% of training.
    Shows how much the policy improved.
    """
    metrics = [
        ("episode/failures",           "Failures/Ep",    "lower"),
        ("episode/weighted_tardiness",  "Wt. Tardiness",  "lower"),
        ("episode/jobs_completed",      "Jobs Completed", "higher"),
        ("episode/avg_health",          "Avg Health %",   "higher"),
        ("episode/service_level",       "Service Level",  "higher"),
        ("episode/pm_cm_ratio",         "PM/CM Ratio",    "higher"),
    ]

    early_means, late_means, labels, directions = [], [], [], []

    for tag, label, direction in metrics:
        if tag not in data:
            continue
        vals = [v for _, v in data[tag]]
        if len(vals) < 10:
            continue
        n = len(vals)
        cutoff = max(n // 5, 5)
        early_means.append(np.mean(vals[:cutoff]))
        late_means.append(np.mean(vals[n - cutoff:]))
        labels.append(label)
        directions.append(direction)

    if not labels:
        print("Not enough data for early vs late comparison")
        return

    x = np.arange(len(labels))
    w = 0.35

    fig, ax = plt.subplots(figsize=(12, 5))
    fig.patch.set_facecolor(COLOURS["bg"])
    ax.set_facecolor(COLOURS["grid"])
    ax.tick_params(colors=COLOURS["text"])
    ax.xaxis.label.set_color(COLOURS["text"])
    ax.yaxis.label.set_color(COLOURS["text"])
    ax.title.set_color(COLOURS["text"])
    for spine in ax.spines.values():
        spine.set_edgecolor(COLOURS["grid"])

    bars_early = ax.bar(x - w/2, early_means, w, label="Early Training (first 20%)",
                        color=COLOURS["r1"], alpha=0.85)
    bars_late  = ax.bar(x + w/2, late_means,  w, label="Late Training (last 20%)",
                        color=COLOURS["r2"], alpha=0.85)

    # Improvement arrows
    for i, (e, l, d) in enumerate(zip(early_means, late_means, directions)):
        improved = (l < e) if d == "lower" else (l > e)
        symbol = "▲" if improved else "▼"
        col    = "#00DC6E" if improved else "#FF1E37"
        pct    = abs((l - e) / max(abs(e), 1e-6)) * 100
        ax.text(x[i], max(e, l) * 1.05,
                f"{symbol}{pct:.0f}%", ha="center", va="bottom",
                fontsize=8, color=col, fontweight="bold")

    ax.set_xticks(x)
    ax.set_xticklabels(labels, fontsize=9, color=COLOURS["text"])
    ax.legend(fontsize=9, facecolor=COLOURS["grid"], labelcolor=COLOURS["text"])
    ax.set_title("Early vs Late Training Comparison",
                 fontsize=12, fontweight="bold", color=COLOURS["text"])
    ax.grid(True, axis="y", color="#2A2E3A", linewidth=0.5)

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    plt.savefig(output_path, dpi=150, bbox_inches="tight", facecolor=COLOURS["bg"])
    plt.close()
    print(f"Saved: {output_path}")


# ---------------------------------------------------------------------------
# EXCEL EXPORT
# ---------------------------------------------------------------------------
def make_training_excel(data: dict, output_path: str) -> None:
    """
    Generates Excel report with 3 sheets:
      Sheet 1: Raw episode data (one row per episode)
      Sheet 2: Early vs Late comparison table
      Sheet 3: Summary statistics (mean, std, min, max, trend)
    """
    if not EXCEL_AVAILABLE:
        print("openpyxl not installed — skipping Excel export")
        return

    wb = openpyxl.Workbook()

    # ---- Styles ----
    hdr_fill   = PatternFill("solid", fgColor="1A1E28")
    hdr_font   = Font(bold=True, color="00DC6E", size=10)
    data_font  = Font(color="D2DCF7", size=9)
    good_fill  = PatternFill("solid", fgColor="0D2B1A")
    bad_fill   = PatternFill("solid", fgColor="2B0D0D")
    center     = Alignment(horizontal="center")
    thin = Border(
        left=Side(style="thin", color="2A2E3A"),
        right=Side(style="thin", color="2A2E3A"),
        top=Side(style="thin", color="2A2E3A"),
        bottom=Side(style="thin", color="2A2E3A"),
    )

    def style_header(cell):
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = thin

    def style_data(cell, good=None):
        cell.font = data_font
        cell.alignment = center
        cell.border = thin
        if good is True:
            cell.fill = good_fill
        elif good is False:
            cell.fill = bad_fill

    # ---- SHEET 1: Raw Episode Data ----
    ws1 = wb.active
    ws1.title = "Episode Data"
    ws1.sheet_properties.tabColor = "00DC6E"

    episode_tags = [
        ("episode/return1",           "Return Agent1"),
        ("episode/return2",           "Return Agent2"),
        ("episode/failures",          "Failures"),
        ("episode/n_PM",              "PM Events"),
        ("episode/n_CM",              "CM Events"),
        ("episode/pm_cm_ratio",       "PM/CM Ratio"),
        ("episode/jobs_completed",    "Jobs Completed"),
        ("episode/weighted_tardiness","Wt. Tardiness"),
        ("episode/service_level",     "Service Level"),
        ("episode/avg_health",        "Avg Health %"),
        ("episode/MTBF",              "MTBF"),
        ("episode/avg_inventory",     "Avg Inventory"),
    ]

    # Find which tags have data
    available = [(tag, label) for tag, label in episode_tags if tag in data]

    headers = ["Episode"] + [label for _, label in available]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        style_header(cell)
        ws1.column_dimensions[get_column_letter(col)].width = 16

    # Find max episodes
    max_ep = max(
        (max(ep for ep, _ in data[tag]) for tag, _ in available if tag in data),
        default=0
    )

    # Build per-episode lookup
    ep_data = {tag: dict(data[tag]) for tag, _ in available if tag in data}

    for ep in range(max_ep + 1):
        row = [ep] + [ep_data[tag].get(ep, "") for tag, _ in available]
        for col, val in enumerate(row, 1):
            cell = ws1.cell(row=ep + 2, column=col, value=val if val != "" else None)
            style_data(cell)

    # Freeze header row
    ws1.freeze_panes = "A2"

    # ---- SHEET 2: Early vs Late ----
    ws2 = wb.create_sheet("Early vs Late")
    ws2.sheet_properties.tabColor = "00C8FF"

    ws2.cell(row=1, column=1, value="BTP2 MARL — Early vs Late Training Comparison")
    ws2["A1"].font = Font(bold=True, color="00DC6E", size=12)
    ws2.merge_cells("A1:G1")

    early_late_headers = ["Metric", "Early Mean", "Early Std", "Late Mean", "Late Std",
                          "Change %", "Improved?"]
    for col, h in enumerate(early_late_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        style_header(cell)
        ws2.column_dimensions[get_column_letter(col)].width = 18

    comparison_tags = [
        ("episode/return1",           "Agent 1 Return",   "higher"),
        ("episode/return2",           "Agent 2 Return",   "higher"),
        ("episode/failures",          "Failures/Ep",      "lower"),
        ("episode/weighted_tardiness","Wt. Tardiness",     "lower"),
        ("episode/jobs_completed",    "Jobs Completed",    "higher"),
        ("episode/service_level",     "Service Level",     "higher"),
        ("episode/avg_health",        "Avg Health %",      "higher"),
        ("episode/pm_cm_ratio",       "PM/CM Ratio",       "higher"),
        ("train/entropy1",            "Entropy Agent 1",   "stable"),
        ("train/entropy2",            "Entropy Agent 2",   "stable"),
    ]

    for row_i, (tag, label, direction) in enumerate(comparison_tags, 4):
        if tag not in data:
            continue
        vals = [v for _, v in data[tag]]
        if len(vals) < 10:
            continue
        n = len(vals)
        cutoff = max(n // 5, 5)
        early = vals[:cutoff]
        late  = vals[n - cutoff:]

        e_mean, e_std = np.mean(early), np.std(early)
        l_mean, l_std = np.mean(late),  np.std(late)
        pct_change = (l_mean - e_mean) / max(abs(e_mean), 1e-6) * 100

        if direction == "higher":
            improved = l_mean > e_mean
        elif direction == "lower":
            improved = l_mean < e_mean
        else:
            improved = abs(pct_change) < 10  # stable within 10%

        row_data = [label, round(e_mean, 3), round(e_std, 3),
                    round(l_mean, 3), round(l_std, 3),
                    f"{pct_change:+.1f}%", "✓ YES" if improved else "✗ NO"]

        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_i, column=col, value=val)
            if col >= 6:
                style_data(cell, good=improved)
            else:
                style_data(cell)

    # ---- SHEET 3: Summary Statistics ----
    ws3 = wb.create_sheet("Summary Stats")
    ws3.sheet_properties.tabColor = "FFB900"

    ws3.cell(row=1, column=1, value="BTP2 MARL — Training Summary Statistics")
    ws3["A1"].font = Font(bold=True, color="FFB900", size=12)
    ws3.merge_cells("A1:G1")

    stat_headers = ["Metric", "N Points", "Mean", "Std", "Min", "Max",
                    "Trend (slope/1000 steps)"]
    for col, h in enumerate(stat_headers, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        style_header(cell)
        ws3.column_dimensions[get_column_letter(col)].width = 22

    all_tags = (
        [(t, l) for t, l, _ in TAGS["episode"]] +
        [(t, l) for t, l, _ in TAGS["training"]]
    )

    for row_i, (tag, label) in enumerate(all_tags, 4):
        if tag not in data:
            continue
        steps_arr = np.array([s for s, _ in data[tag]], dtype=float)
        vals_arr  = np.array([v for _, v in data[tag]], dtype=float)
        n = len(vals_arr)
        if n < 2:
            continue

        # Linear trend per 1000 steps
        if steps_arr[-1] > steps_arr[0]:
            slope = np.polyfit(steps_arr / 1000.0, vals_arr, 1)[0]
        else:
            slope = 0.0

        row_data = [label, n, round(float(np.mean(vals_arr)), 4),
                    round(float(np.std(vals_arr)), 4),
                    round(float(np.min(vals_arr)), 4),
                    round(float(np.max(vals_arr)), 4),
                    round(slope, 6)]

        for col, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_i, column=col, value=val)
            style_data(cell)

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    print(f"Saved: {output_path}")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--logdir",   default="runs/",      help="TensorBoard log directory")
    parser.add_argument("--outdir",   default="results/",   help="Output directory")
    parser.add_argument("--no-excel", action="store_true",  help="Skip Excel export")
    parser.add_argument("--no-plots", action="store_true",  help="Skip PNG plots")
    args = parser.parse_args()

    os.makedirs(args.outdir, exist_ok=True)
    print(f"\nLoading TensorBoard data from: {args.logdir}")
    data = load_tb_data(args.logdir)

    if not data:
        print("No data found. Run some training first:")
        print("  python scripts/train.py --config configs/phase1.yaml --timesteps 5000")
        return

    if not args.no_plots:
        print("\nGenerating training curve plots...")
        make_training_plots(data, os.path.join(args.outdir, "training_curves.png"))
        make_early_vs_late_plot(data, os.path.join(args.outdir, "early_vs_late.png"))

    if not args.no_excel:
        print("\nGenerating Excel report...")
        make_training_excel(data, os.path.join(args.outdir, "training_analysis.xlsx"))

    print(f"\nAll outputs in: {args.outdir}/")
    print("  training_curves.png   — 4-panel training curves")
    print("  early_vs_late.png     — before/after comparison bar chart")
    print("  training_analysis.xlsx — 3-sheet Excel report")


if __name__ == "__main__":
    main()
