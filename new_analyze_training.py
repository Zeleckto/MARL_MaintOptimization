"""
analyze_training.py — Training Curve Analysis
==============================================
Reads TensorBoard data and generates conference-quality training analysis:
  - Training curves (rewards, losses, KPIs over time)
  - Early vs Late stage comparison table
  - Convergence metrics (steps to availability>0.8, entropy decay)
  - Resource dynamics over training (renewable/consumable from debug/ tags)
  - Reward signal decomposition (was dense noise dominating?)

Usage:
    python analyze_training.py                        # reads outputs/runs/
    python analyze_training.py --logdir outputs/runs/phase1_TIMESTAMP

Output → outputs/results/training/
    training_curves.png
    resource_dynamics.png
    convergence_report.txt
    training_analysis.xlsx
"""
import argparse, os, sys, glob
import numpy as np
from collections import defaultdict

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)
sys.path.insert(0, os.path.join(ROOT, "analytics"))

try:
    from tensorboard.backend.event_processing.event_accumulator import EventAccumulator
    TB_OK = True
except ImportError:
    TB_OK = False

import matplotlib; matplotlib.use("Agg")
import matplotlib.pyplot as plt
from analytics.plot_utils import BG, AXES, TEXT, MUTED, MARL_C, GOLD_C, PALETTE, _style_ax


def load_tb(logdirs):
    raw = defaultdict(dict)
    n_runs = 0
    for ld in logdirs:
        for run_dir in (sorted(os.listdir(ld)) if os.path.isdir(ld) else [ld]):
            path = os.path.join(ld, run_dir) if os.path.isdir(ld) else ld
            if not os.path.isdir(path): continue
            try:
                ea = EventAccumulator(path, size_guidance={EventAccumulator.SCALARS: 0})
                ea.Reload()
                tags = ea.Tags().get("scalars", [])
                if not tags: continue
                n_runs += 1
                for tag in tags:
                    for e in ea.Scalars(tag):
                        raw[tag][e.step] = e.value
            except: continue
    data = {t: {"steps": np.array(sorted(sv)), "values": np.array([sv[s] for s in sorted(sv)])}
            for t, sv in raw.items()}
    print(f"  Loaded {n_runs} run(s) | {len(data)} tags")
    return data


def smooth(arr, w=None):
    if len(arr) == 0: return arr
    w = w or max(len(arr)//25, 1)
    return np.convolve(arr, np.ones(w)/w, mode="valid")


def make_training_curves(data, outpath):
    """8-panel training progress figure."""
    panels = [
        ("episode/availability",    "Availability",       MARL_C,  "UP"),
        ("episode/failures",        "Failures / ep",      "#FF3246","DN"),
        ("train/critic_loss",       "Critic Loss",        GOLD_C,  "DN"),
        ("train/entropy1",          "Entropy A1",         "#B43CFF","DN"),
        ("rewards/agent1_r1",       "r₁ / step",          MARL_C,  "UP"),
        ("rewards/agent2_r2",       "r₂ / step",          "#00DC6E","UP"),
        ("rewards/shared",          "r_shared / step",    "#FFB900","DN"),
        ("episode/n_PM",            "PM Events / ep",     "#00DC6E","UP"),
    ]
    present = [(t, l, c, d) for t, l, c, d in panels if t in data]
    if not present: print("  No data for training curves."); return

    n = len(present); cols = 4; rows = (n + cols - 1)//cols
    fig, axes = plt.subplots(rows, cols, figsize=(cols*4, rows*3.2))
    fig.patch.set_facecolor(BG)
    axes_flat = np.array(axes).flatten()

    for i, (tag, label, colour, direction) in enumerate(present):
        ax = axes_flat[i]
        d  = data[tag]
        x, y = d["steps"], d["values"]
        y_sm = smooth(y); x_sm = x[:len(y_sm)]
        ax.plot(x, y, alpha=0.15, color=colour, linewidth=0.7)
        ax.plot(x_sm, y_sm, color=colour, linewidth=2.0, label="smoothed")
        ax.axhline(0, color="#2A2D35", linewidth=0.7, linestyle="--")
        _style_ax(ax, title=label, xlabel="Training step")
        # Add target line for key metrics
        if tag == "episode/availability":
            ax.axhline(0.85, color=GOLD_C, linewidth=1, linestyle=":", alpha=0.8)
            ax.text(x[-1]*0.02, 0.86, "Target 0.85", color=GOLD_C, fontsize=7)
        if tag == "train/entropy1":
            import math
            max_ent = 5 * math.log(2)
            ax.axhline(max_ent, color="#FF3246", linewidth=1, linestyle=":", alpha=0.7)
            ax.text(x[-1]*0.02, max_ent*1.01, f"Max ({max_ent:.2f})", color="#FF3246", fontsize=7)

    for j in range(i+1, len(axes_flat)):
        axes_flat[j].set_visible(False)

    fig.suptitle("Training Progress", color=TEXT, size=13, fontweight="bold", y=1.02)
    plt.tight_layout()
    os.makedirs(os.path.dirname(outpath) if os.path.dirname(outpath) else ".", exist_ok=True)
    fig.savefig(outpath, dpi=200, bbox_inches="tight", facecolor=BG)
    plt.close(fig)
    print(f"  Saved: {outpath}")


def make_resource_plot(data, outpath):
    """Resource dynamics over training — shows ordering and depletion."""
    tags = {
        "debug/renewable_0": ("Technicians (K=3)",   "#00C8FF"),
        "debug/renewable_2": ("Maint. Bays (K=4)",   "#B43CFF"),
        "debug/consumable_0":("Spare Parts",          "#00DC6E"),
        "debug/consumable_2":("Consumable Tools",     "#FF3246"),
    }
    present = [(t, l, c) for t, (l, c) in tags.items() if t in data]
    if not present: print("  No debug/ resource tags in TB data."); return

    fig, axes = plt.subplots(1, 2, figsize=(10, 4))
    fig.patch.set_facecolor(BG)

    ren_tags = [(t,l,c) for t,l,c in present if "renewable" in t]
    con_tags = [(t,l,c) for t,l,c in present if "consumable" in t]

    for ax, tag_list, title in [(axes[0], ren_tags, "Renewable Resources"),
                                  (axes[1], con_tags, "Consumable Inventory")]:
        for tag, label, colour in tag_list:
            d  = data[tag]
            y_sm = smooth(d["values"]); x_sm = d["steps"][:len(y_sm)]
            ax.plot(x_sm, y_sm, color=colour, linewidth=1.8, label=label)
            ax.plot(d["steps"], d["values"], alpha=0.12, color=colour, linewidth=0.6)
        _style_ax(ax, title=title, xlabel="Step", ylabel="Count / Units")
        ax.legend(facecolor=AXES, edgecolor=MUTED, labelcolor=TEXT, fontsize=8)

    fig.suptitle("Resource Dynamics During Training", color=TEXT, size=12, fontweight="bold")
    plt.tight_layout()
    os.makedirs(os.path.dirname(outpath) if os.path.dirname(outpath) else ".", exist_ok=True)
    fig.savefig(outpath, dpi=200, bbox_inches="tight", facecolor=BG)
    plt.close(fig)
    print(f"  Saved: {outpath}")


def make_reward_decomposition(data, outpath):
    """Bar chart comparing reward component magnitudes — dense vs sparse signal."""
    tags_to_check = {
        "rewards/agent1_r1":  ("r₁ total",       MARL_C),
        "rewards/agent2_r2":  ("r₂ total",       "#00DC6E"),
        "rewards/shared":     ("r_shared",        GOLD_C),
    }
    present = [(l, c, np.mean(data[t]["values"]))
               for t, (l, c) in tags_to_check.items() if t in data]
    if not present: return

    fig, ax = plt.subplots(figsize=(7, 4)); fig.patch.set_facecolor(BG)
    labels  = [p[0] for p in present]
    vals    = [abs(p[2]) for p in present]
    colours = [p[1] for p in present]
    bars = ax.bar(range(len(labels)), vals, color=colours, alpha=0.88,
                  edgecolor="#2A2D35", linewidth=0.8)
    ax.set_xticks(range(len(labels))); ax.set_xticklabels(labels, color=MUTED)
    _style_ax(ax, title="Mean |Reward| per Step — Signal Magnitude",
              ylabel="|mean value|")
    for bar, p in zip(bars, present):
        ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.003,
                f"{p[2]:.4f}", ha="center", va="bottom", color=TEXT, fontsize=8)
    plt.tight_layout()
    os.makedirs(os.path.dirname(outpath) if os.path.dirname(outpath) else ".", exist_ok=True)
    fig.savefig(outpath, dpi=200, bbox_inches="tight", facecolor=BG)
    plt.close(fig)
    print(f"  Saved: {outpath}")


def early_vs_late(data, n_pct=20):
    """Compare first n_pct% vs last n_pct% of training for key tags."""
    results = {}
    for tag, d in data.items():
        v = d["values"]
        if len(v) < 10: continue
        cut = max(1, len(v) * n_pct // 100)
        early = v[:cut]; late = v[-cut:]
        results[tag] = {
            "early_mean": float(np.mean(early)), "early_std": float(np.std(early)),
            "late_mean":  float(np.mean(late)),  "late_std":  float(np.std(late)),
            "improvement": float(np.mean(late) - np.mean(early)),
            "improvement_pct": float((np.mean(late)-np.mean(early)) / max(abs(np.mean(early)), 1e-8)*100),
            "n_points": len(v),
        }
    return results


def convergence_report(data, outpath):
    """Write a convergence analysis text report."""
    import math
    lines = ["CONVERGENCE ANALYSIS REPORT", "="*50]
    max_ent_2  = 5 * math.log(2)

    # Availability
    if "episode/availability" in data:
        v = data["episode/availability"]["values"]
        s = data["episode/availability"]["steps"]
        for tgt in [0.70, 0.80, 0.85]:
            idx = next((i for i, x in enumerate(v) if x >= tgt), None)
            lines.append(f"Steps to availability>{tgt}: "
                         + (f"{s[idx]:,}" if idx else "NOT REACHED"))
        lines.append(f"Final availability: {v[-1]:.4f}")

    # Entropy decay
    if "train/entropy1" in data:
        e = data["train/entropy1"]["values"]
        s = data["train/entropy1"]["steps"]
        lines.append(f"\nEntropy: start={e[0]:.4f}  end={e[-1]:.4f}  max={max_ent_2:.4f}")
        tgt_e = max_ent_2 * 0.85
        idx   = next((i for i, x in enumerate(e) if x < tgt_e), None)
        lines.append(f"Steps to entropy < {tgt_e:.3f} (15% drop): "
                     + (f"{s[idx]:,}" if idx else "NOT REACHED"))
        if abs(e[-1] - max_ent_2) < 0.05:
            lines.append("WARNING: Entropy still at max — policy may not have specialised")

    # Critic loss
    if "train/critic_loss" in data:
        c = data["train/critic_loss"]["values"]
        lines.append(f"\nCritic loss: start={c[0]:.4f}  end={c[-1]:.4f}")
        if np.all(c < 1e-6):
            lines.append("WARNING: Critic loss = 0 throughout — critic not training!")
        else:
            lines.append("Critic loss is non-zero — critic trained correctly")

    # Failures
    if "episode/failures" in data:
        v = data["episode/failures"]["values"]
        lines.append(f"\nFailures: start={v[0]:.2f}  end={v[-1]:.2f}  "
                     f"reduction={v[0]-v[-1]:+.2f}")

    # r_shared check
    if "rewards/shared" in data:
        rs = data["rewards/shared"]["values"]
        pct_nz = (rs != 0).mean() * 100
        lines.append(f"\nrewards/shared: {pct_nz:.1f}% of steps non-zero")
        if pct_nz < 0.1:
            lines.append("WARNING: r_shared always 0 — Bug 3 may not be fixed!")

    os.makedirs(os.path.dirname(outpath) if os.path.dirname(outpath) else ".", exist_ok=True)
    with open(outpath, "w") as f: f.write("\n".join(lines))
    for l in lines: print(f"    {l}")
    print(f"  Saved: {outpath}")


def write_excel(data, el_table, outpath):
    """Write training analysis Excel workbook."""
    try:
        import openpyxl
        from analytics.excel_writer import _cell, HEADER, BG, TEXT_FG, GOLD, SUB_FG, MARL_BG, MARL_FG
        wb = openpyxl.Workbook()

        # Sheet 1: Early vs Late
        ws = wb.active; ws.title = "Early vs Late"
        ws.sheet_view.showGridLines = False
        headers = ["Tag","n_points","Early mean","Early std","Late mean","Late std","Δ","Δ%"]
        for ci, h in enumerate(headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 18
            _cell(ws, 1, ci, h, HEADER, GOLD, bold=True)
        for ri, (tag, ev) in enumerate(sorted(el_table.items()), 2):
            vals = [tag, ev["n_points"],
                    round(ev["early_mean"],4), round(ev["early_std"],4),
                    round(ev["late_mean"],4),  round(ev["late_std"],4),
                    round(ev["improvement"],4), round(ev["improvement_pct"],1)]
            for ci, v in enumerate(vals, 1):
                fg = MARL_FG if ci == 8 and ev["improvement_pct"] > 0 else TEXT_FG
                _cell(ws, ri, ci, v, BG, fg)

        # Sheet 2: Key TB values summary
        ws2 = wb.create_sheet("TB Summary")
        ws2.sheet_view.showGridLines = False
        _cell(ws2, 1, 1, "Tag", HEADER, GOLD, bold=True)
        for ci, h in enumerate(["n_pts","first","last","min","max","mean"], 2):
            _cell(ws2, 1, ci, h, HEADER, GOLD, bold=True)
        for ri, (tag, d) in enumerate(sorted(data.items()), 2):
            v = d["values"]
            vals = [tag, len(v), round(float(v[0]),4), round(float(v[-1]),4),
                    round(float(v.min()),4), round(float(v.max()),4), round(float(v.mean()),4)]
            for ci, val in enumerate(vals, 1):
                _cell(ws2, ri, ci, val, BG, TEXT_FG)

        os.makedirs(os.path.dirname(outpath) if os.path.dirname(outpath) else ".", exist_ok=True)
        wb.save(outpath)
        print(f"  Saved: {outpath}")
    except Exception as ex:
        print(f"  Excel skipped: {ex}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--logdir", nargs="+", default=["outputs/runs/"])
    ap.add_argument("--outdir", default="outputs/results/training/")
    ap.add_argument("--early-pct", type=int, default=20)
    args = ap.parse_args()

    if not TB_OK:
        print("pip install tensorboard"); sys.exit(1)

    os.makedirs(args.outdir, exist_ok=True)
    print(f"\n{'='*55}")
    print(f"  TRAINING ANALYSIS")
    print(f"{'='*55}")

    data = load_tb(args.logdir)
    if not data:
        print("  No TensorBoard data found."); return

    # Training curves
    make_training_curves(data, os.path.join(args.outdir, "training_curves.png"))

    # Resource dynamics
    make_resource_plot(data, os.path.join(args.outdir, "resource_dynamics.png"))

    # Reward decomposition
    make_reward_decomposition(data, os.path.join(args.outdir, "reward_decomposition.png"))

    # Convergence report
    convergence_report(data, os.path.join(args.outdir, "convergence_report.txt"))

    # Early vs late
    el = early_vs_late(data, n_pct=args.early_pct)

    # Excel
    write_excel(data, el, os.path.join(args.outdir, "training_analysis.xlsx"))

    print(f"\n  All outputs: {args.outdir}\n")


if __name__ == "__main__":
    main()