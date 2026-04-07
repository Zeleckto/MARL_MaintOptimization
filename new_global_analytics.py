"""
global_analytics.py — Master Post-Training Analytics
=====================================================
Runs EVERYTHING in one command after Phase 3 training is done.

Usage:
    python global_analytics.py --phase3-ckpt outputs/checkpoints/phase3_end.pt
    python global_analytics.py --phase3-ckpt ... --episodes 100 --outdir outputs/results/paper/

Generates:
    paper/baselines/            → comparison tables + charts (all 3 stoch levels)
    paper/training/             → training curves, convergence analysis
    paper/ablations/            → all ablation study results
    paper/stats/                → significance tests
    paper/lambda_sensitivity/   → λ sensitivity analysis (from TB data)
    paper/convergence/          → convergence speed comparison
    paper/MASTER_TABLE.xlsx     → all paper tables in one workbook
"""
import argparse, os, sys, yaml, shutil
import numpy as np

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)
sys.path.insert(0, os.path.join(ROOT, "analytics"))


def run_baselines(phase3_ckpt, episodes, outdir):
    """Run baseline comparison for all 3 stochasticity levels."""
    for stoch in [1, 2, 3]:
        print(f"\n{'='*55}")
        print(f"  BASELINES — stoch_level={stoch}")
        print(f"{'='*55}")
        d = os.path.join(outdir, f"baselines/stoch{stoch}/")
        os.makedirs(d, exist_ok=True)
        cmd = [sys.executable, "analyze_baselines.py",
               "--episodes", str(episodes),
               "--stoch-level", str(stoch),
               "--outdir", d]
        if phase3_ckpt:
            cmd += ["--checkpoint", phase3_ckpt]
        import subprocess
        subprocess.run(cmd, check=False)


def run_training_analysis(outdir):
    """Extract training curves from TensorBoard."""
    print(f"\n{'='*55}")
    print(f"  TRAINING ANALYSIS")
    print(f"{'='*55}")
    d = os.path.join(outdir, "training/")
    os.makedirs(d, exist_ok=True)
    try:
        from tensorboard.backend.event_processing import event_accumulator
        tb_base = "outputs/runs/"
        if not os.path.exists(tb_base):
            print("  No TB data found in outputs/runs/"); return

        raw = {}
        for run_dir in sorted(os.listdir(tb_base)):
            path = os.path.join(tb_base, run_dir)
            if not os.path.isdir(path): continue
            try:
                ea = event_accumulator.EventAccumulator(path,
                     size_guidance={event_accumulator.SCALARS: 0})
                ea.Reload()
                for tag in ea.Tags().get("scalars", []):
                    if tag not in raw: raw[tag] = {}
                    for e in ea.Scalars(tag):
                        raw[tag][e.step] = e.value
            except: continue

        data = {t: {"steps": np.array(sorted(sv)),
                    "values": np.array([sv[s] for s in sorted(sv)])}
                for t, sv in raw.items()}
        print(f"  Loaded {len(data)} TB tags")

        from analytics.plot_utils import training_curves
        training_curves(data,
                        outpath=os.path.join(d, "training_curves.png"),
                        title="MARL Training Progress")

        # Convergence speed analysis
        _convergence_analysis(data, d)

        # Lambda sensitivity from TB if available
        _lambda_sensitivity(data, os.path.join(outdir, "lambda_sensitivity/"))

        # Write training summary Excel
        _training_excel(data, os.path.join(d, "training_summary.xlsx"))

    except ImportError:
        print("  tensorboard not installed — skipping")
    except Exception as ex:
        print(f"  Training analysis error: {ex}")


def _convergence_analysis(data, outdir):
    """Analyse convergence speed: steps to 80% availability, steps to critic convergence."""
    os.makedirs(outdir, exist_ok=True)
    report = []
    if "episode/availability" in data:
        avail  = data["episode/availability"]["values"]
        steps  = data["episode/availability"]["steps"]
        tgt    = 0.80
        idx    = next((i for i, v in enumerate(avail) if v >= tgt), None)
        report.append(f"Steps to availability>{tgt}: {steps[idx]:,}" if idx else
                      f"Availability never reached {tgt}")
    if "train/critic_loss" in data:
        cl    = data["train/critic_loss"]["values"]
        cl_steps = data["train/critic_loss"]["steps"]
        if np.any(cl > 0):
            first_nz = next((i for i, v in enumerate(cl) if v > 0), None)
            report.append(f"Critic loss first non-zero at step: {cl_steps[first_nz]:,}" if first_nz else "Critic loss always 0!")
    if "train/entropy1" in data:
        ent = data["train/entropy1"]["values"]
        ent_steps = data["train/entropy1"]["steps"]
        import math
        max_ent = 5 * math.log(2)
        tgt_ent = max_ent * 0.8  # 20% reduction
        idx = next((i for i, v in enumerate(ent) if v < tgt_ent), None)
        report.append(f"Entropy fell 20% (below {tgt_ent:.3f}) at step: {ent_steps[idx]:,}" if idx else
                      "Entropy did not drop 20%")

    print("  CONVERGENCE:")
    for r in report: print(f"    {r}")
    with open(os.path.join(outdir, "convergence_report.txt"), "w") as f:
        f.write("\n".join(report))


def _lambda_sensitivity(data, outdir):
    """If multiple runs with different lambda exist, plot sensitivity."""
    # This would need multiple runs — placeholder that checks for tag
    os.makedirs(outdir, exist_ok=True)
    # Check if Ablations/A12 results exist
    a12_dir = "outputs/results/ablations/A12_lambda/"
    if os.path.exists(a12_dir):
        shutil.copytree(a12_dir, os.path.join(outdir, "A12_results"), dirs_exist_ok=True)
        print(f"  Lambda sensitivity results copied from {a12_dir}")
    else:
        print(f"  Run Ablations/A12_lambda_sensitivity.py first for λ analysis")


def _training_excel(data, outpath):
    """Write training KPI summary to Excel."""
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Training Summary"
        key_tags = ["episode/availability","episode/failures","episode/n_PM",
                    "episode/n_CM","train/critic_loss","train/entropy1",
                    "rewards/agent1_r1","rewards/shared"]
        ws.append(["Tag","n_points","first_val","last_val","min","max","mean"])
        for tag in key_tags:
            if tag not in data: continue
            v = data[tag]["values"]
            ws.append([tag, len(v), round(float(v[0]),4), round(float(v[-1]),4),
                       round(float(v.min()),4), round(float(v.max()),4),
                       round(float(v.mean()),4)])
        os.makedirs(os.path.dirname(outpath), exist_ok=True)
        wb.save(outpath)
        print(f"  Saved: {outpath}")
    except Exception as ex:
        print(f"  Training Excel skipped: {ex}")


def run_ablations(phase3_ckpt, episodes, outdir):
    """Run all ablation studies."""
    print(f"\n{'='*55}")
    print(f"  ABLATION STUDIES")
    print(f"{'='*55}")
    ablation_dir = os.path.join(outdir, "ablations/")
    import subprocess
    cmd = [sys.executable, "Ablations/run_all_ablations.py",
           "--group", "1",
           "--outdir", ablation_dir]
    if phase3_ckpt:
        cmd += ["--checkpoint", phase3_ckpt]
    subprocess.run(cmd, check=False)


def run_stats(csv_path, outdir):
    """Run statistical comparison."""
    print(f"\n{'='*55}")
    print(f"  STATISTICAL ANALYSIS")
    print(f"{'='*55}")
    if not os.path.exists(csv_path):
        print(f"  CSV not found: {csv_path}"); return
    d = os.path.join(outdir, "stats/")
    import subprocess
    subprocess.run([sys.executable, "compare_results.py",
                    "--csv", csv_path, "--outdir", d], check=False)


def build_master_table(outdir):
    """Aggregate all results into one MASTER_TABLE.xlsx workbook."""
    print(f"\n{'='*55}")
    print(f"  BUILDING MASTER TABLE")
    print(f"{'='*55}")
    try:
        import openpyxl
        wb_master = openpyxl.Workbook()
        wb_master.remove(wb_master.active)

        # Copy sheets from each generated workbook
        sources = [
            (os.path.join(outdir, "baselines/stoch3/comparison_table.xlsx"), "Baselines (stoch=3)"),
            (os.path.join(outdir, "baselines/stoch1/comparison_table.xlsx"), "Baselines (stoch=1)"),
            (os.path.join(outdir, "stats/significance_table.xlsx"),          "Statistical Tests"),
        ]
        for src_path, sheet_label in sources:
            if not os.path.exists(src_path): continue
            wb_src = openpyxl.load_workbook(src_path)
            for sname in wb_src.sheetnames:
                ws_src  = wb_src[sname]
                ws_dest = wb_master.create_sheet(f"{sheet_label[:15]}/{sname[:15]}")
                for row in ws_src:
                    for cell in row:
                        nc = ws_dest.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            nc.fill      = cell.fill
                            nc.font      = cell.font
                            nc.alignment = cell.alignment

        path = os.path.join(outdir, "MASTER_TABLE.xlsx")
        wb_master.save(path)
        print(f"  MASTER TABLE: {path}")
    except Exception as ex:
        print(f"  Master table skipped: {ex}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--phase3-ckpt", default=None)
    ap.add_argument("--phase1-ckpt", default=None)
    ap.add_argument("--phase2-ckpt", default=None)
    ap.add_argument("--episodes",    type=int, default=50)
    ap.add_argument("--outdir",      default="outputs/results/paper/")
    ap.add_argument("--skip-baselines",  action="store_true")
    ap.add_argument("--skip-training",   action="store_true")
    ap.add_argument("--skip-ablations",  action="store_true")
    ap.add_argument("--skip-stats",      action="store_true")
    args = ap.parse_args()

    os.makedirs(args.outdir, exist_ok=True)
    print(f"\n{'='*60}")
    print(f"  GLOBAL ANALYTICS — BTP2 MARL Manufacturing")
    print(f"  Output: {args.outdir}")
    print(f"{'='*60}")

    # Step 1: Baselines
    if not args.skip_baselines:
        run_baselines(args.phase3_ckpt, args.episodes, args.outdir)

    # Step 2: Training analysis
    if not args.skip_training:
        run_training_analysis(args.outdir)

    # Step 3: Ablations
    if not args.skip_ablations and args.phase3_ckpt:
        run_ablations(args.phase3_ckpt, args.episodes, args.outdir)
    elif not args.phase3_ckpt:
        print("\n  Ablations skipped — no checkpoint provided")

    # Step 4: Statistical comparison
    if not args.skip_stats:
        csv3 = os.path.join(args.outdir, "baselines/stoch3/episode_data.csv")
        run_stats(csv3, args.outdir)

    # Step 5: Master table
    build_master_table(args.outdir)

    print(f"\n{'='*60}")
    print(f"  DONE — All outputs in {args.outdir}")
    print(f"  Key files:")
    for root, dirs, files in os.walk(args.outdir):
        for f in files:
            if f.endswith((".xlsx",".png")):
                rel = os.path.relpath(os.path.join(root,f), args.outdir)
                print(f"    {rel}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()