"""
run_baselines.py
=================
Runs all 4 baselines + MARL proxy, saves EVERYTHING in structured output folder.

Usage:
    python run_baselines.py                           # 30 episodes
    python run_baselines.py --episodes 50
    python run_baselines.py --outdir results/phase1

Output:
    outputs/baselines/
    +-- summary_table.xlsx / .csv
    +-- per_episode_data.csv
    +-- comparison_bar_chart.png
    +-- radar_chart.png
    +-- capacity_breakdown.png
    +-- kpi_distributions.png
    +-- run_metadata.json
    +-- per_baseline/*.csv
"""
import argparse, os, sys, csv, json
import numpy as np
from datetime import datetime
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import yaml
from environments.mfg_env import ManufacturingEnv, AGENT_PDM, AGENT_JOBSHOP
from environments.transitions.degradation import MachineStatus
from environments.transitions.job_dynamics import OpStatus
from benchmarks.baselines import get_all_baselines
import matplotlib; matplotlib.use("Agg")
import matplotlib.pyplot as plt

class MARLProxyBaseline:
    name = "MARL_proxy(h<75)"
    def reset(self): pass
    def agent1_action(self, env):
        maint = np.zeros(len(env.machine_states), dtype=int)
        for i, s in enumerate(env.machine_states):
            if s.status == MachineStatus.OP and not env.machine_busy[i] and s.health < 75:
                maint[i] = 1
        inv = env.resource_state.consumable_inventory
        reorder = np.array([8.0 if inv[i] < 10 else 0.0 for i in range(len(inv))])
        return {"maintenance": maint, "reorder": reorder}
    def agent2_action(self, env):
        if not env._valid_pairs: return len(env._valid_pairs)
        best = 0; best_score = -999
        for idx, (jid, oidx, mid) in enumerate(env._valid_pairs):
            job = next((j for j in env.jobs if j.job_id == jid), None)
            if not job: continue
            op = job.operations[oidx]
            pt = min(op.nominal_proc_times.values()) / 8.0 if op.nominal_proc_times else 5
            h = env.machine_states[mid].health / 100.0 if mid < len(env.machine_states) else 0.5
            score = -pt + 2*h
            if score > best_score: best_score = score; best = idx
        return best

COLORS = ["#e74c3c", "#f39c12", "#2ecc71", "#3498db", "#9b59b6"]

def run_episode(env, policy, seed):
    env.reset(seed=seed); policy.reset()
    busy=idle=pm_s=cm_s=fail_s=0
    for step in range(env.t_max):
        for i, s in enumerate(env.machine_states):
            if s.status == MachineStatus.OP:
                if env.machine_busy[i]: busy += 1
                else: idle += 1
            elif s.status == MachineStatus.PM: pm_s += 1
            elif s.status == MachineStatus.CM: cm_s += 1
            elif s.status == MachineStatus.FAIL: fail_s += 1
        a1 = policy.agent1_action(env); env._step_agent1(a1)
        a2 = policy.agent2_action(env)
        if env._valid_pairs and isinstance(a2,(int,np.integer)) and a2 < len(env._valid_pairs):
            env._step_agent2(a2)
        else: env._step_agent2(None)
        env._resolve_physics(); env._compute_rewards()
    cj = [j for j in env.jobs if j.completion_time is not None]
    nc = len(cj); not_ = sum(1 for j in cj if j.completion_time <= j.due_date)
    tard = sum(j.weight*max(0,j.completion_time-j.due_date) for j in cj)
    ops_done = sum(1 for j in env.jobs for op in j.operations if op.status==OpStatus.DONE)
    ops_total = sum(len(j.operations) for j in env.jobs)
    ah = float(np.mean([s.health for s in env.machine_states]))
    nop = sum(1 for s in env.machine_states if s.status==MachineStatus.OP)
    mcfg = env.config.get("machines",[])
    w = env.reward_fn.weights
    nf=env._episode_failures; npm=env._episode_pm; ncm=env._episode_cm
    mc = npm*w.get("c_PM",1)+ncm*w.get("c_CM",7)+nf*w.get("c_fail",25)
    oc = env._episode_order_cost; inv=float(env.resource_state.consumable_inventory.sum())
    hc = w.get("w_hold",0.005)*inv*env.t_max
    tc = mc+oc+hc+w.get("alpha",1)*tard
    ms = max((j.completion_time for j in cj),default=float(env.t_max))
    cap = len(env.machine_states)*env.t_max
    mrl = float(np.mean([s.rul/max(m.get("eta",1000),1) for s,m in zip(env.machine_states,mcfg)])) if mcfg else 0
    tms = npm*np.mean([m.get("tau_PM_shifts",2) for m in mcfg])+ncm*np.mean([m.get("tau_CM_shifts",6) for m in mcfg]) if mcfg else 0
    return dict(seed=seed,jobs_completed=nc,ops_completed=ops_done,ops_total=ops_total,
        failures=nf,n_PM=npm,n_CM=ncm,on_time_jobs=not_,late_jobs=nc-not_,
        service_level=not_/max(nc,1),weighted_tardiness=tard,makespan=ms,
        avg_health_end=ah,availability_end=nop/max(len(env.machine_states),1),
        mtbf=float(env.t_max)/max(nf,1),mttr=tms/max(npm+ncm,1),mean_rul_norm=mrl,
        utilization=busy/max(cap,1),busy_shifts=busy,idle_shifts=idle,
        pm_shifts=pm_s,cm_shifts=cm_s,fail_shifts=fail_s,
        maint_cost=mc,order_cost=oc,holding_cost=hc,total_cost=tc,inventory_end=inv,
        return_agent1=env._cumulative_rewards.get(AGENT_PDM,0),
        return_agent2=env._cumulative_rewards.get(AGENT_JOBSHOP,0),
        pm_cm_ratio=npm/max(ncm,1))

def plot_bars(R, outdir):
    kpis=[("jobs_completed","Jobs Completed",True),("failures","Failures",False),
          ("service_level","Service Level",True),("availability_end","Availability",True),
          ("avg_health_end","Health %",True),("weighted_tardiness","Tardiness",False),
          ("n_PM","PM Events",None),("total_cost","Total Cost",False)]
    fig,axes=plt.subplots(2,4,figsize=(20,10)); axes=axes.flatten(); names=list(R.keys())
    for ax,(k,l,hb) in zip(axes,kpis):
        ms=[np.mean([r[k] for r in R[n]]) for n in names]
        ss=[np.std([r[k] for r in R[n]]) for n in names]
        bars=ax.bar(range(len(names)),ms,yerr=ss,color=COLORS[:len(names)],capsize=4,edgecolor="white")
        ax.set_title(l,fontsize=12,fontweight="bold"); ax.set_xticks(range(len(names)))
        ax.set_xticklabels([n.replace("_","\n") for n in names],fontsize=7); ax.grid(axis="y",alpha=0.3)
        if hb is not None:
            bi=np.argmax(ms) if hb else np.argmin(ms); bars[bi].set_edgecolor("gold"); bars[bi].set_linewidth(3)
    plt.tight_layout(); plt.savefig(os.path.join(outdir,"comparison_bar_chart.png"),dpi=150,bbox_inches="tight"); plt.close()

def plot_radar(R, outdir):
    mets=["jobs_completed","service_level","availability_end","avg_health_end","mtbf"]
    inv_m=["failures","weighted_tardiness","total_cost"]
    labels=["Jobs","Svc Lvl","Avail","Health","MTBF","1/Fail","1/Tard","1/Cost"]
    names=list(R.keys()); vals={}
    for n in names:
        row=[np.mean([r[m] for r in R[n]]) for m in mets]+[1/max(np.mean([r[m] for r in R[n]]),0.01) for m in inv_m]
        vals[n]=row
    av=np.array(list(vals.values())); mn=av.min(0); mx=av.max(0); rng=mx-mn; rng[rng==0]=1
    angles=np.linspace(0,2*np.pi,len(labels),endpoint=False).tolist()+[0]
    fig,ax=plt.subplots(figsize=(10,10),subplot_kw=dict(polar=True))
    for i,n in enumerate(names):
        norm=[(v-m)/r for v,m,r in zip(vals[n],mn,rng)]+[((vals[n][0]-mn[0])/rng[0])]
        ax.plot(angles,norm,"o-",color=COLORS[i],linewidth=2,label=n); ax.fill(angles,norm,alpha=0.1,color=COLORS[i])
    ax.set_xticks(angles[:-1]); ax.set_xticklabels(labels,fontsize=10); ax.set_ylim(0,1.1)
    ax.legend(loc="upper right",bbox_to_anchor=(1.3,1.1)); ax.set_title("Baseline Comparison",fontsize=14,fontweight="bold",pad=20)
    plt.savefig(os.path.join(outdir,"radar_chart.png"),dpi=150,bbox_inches="tight"); plt.close()

def plot_capacity(R, outdir):
    names=list(R.keys()); fig,ax=plt.subplots(figsize=(12,6))
    g=lambda n,k: np.mean([r[k] for r in R[n]])
    busy=[g(n,"busy_shifts") for n in names]; idle=[g(n,"idle_shifts") for n in names]
    pm=[g(n,"pm_shifts") for n in names]; cm=[g(n,"cm_shifts") for n in names]
    fl=[g(n,"fail_shifts") for n in names]; x=range(len(names))
    ax.bar(x,busy,label="Busy",color="#2ecc71")
    ax.bar(x,idle,bottom=busy,label="Idle",color="#95a5a6")
    ax.bar(x,pm,bottom=[b+i for b,i in zip(busy,idle)],label="PM",color="#3498db")
    ax.bar(x,cm,bottom=[b+i+p for b,i,p in zip(busy,idle,pm)],label="CM",color="#e74c3c")
    ax.bar(x,fl,bottom=[b+i+p+c for b,i,p,c in zip(busy,idle,pm,cm)],label="Failed",color="#c0392b")
    ax.set_xticks(x); ax.set_xticklabels([n.replace("_","\n") for n in names],fontsize=9)
    ax.set_ylabel("Machine-Shifts"); ax.set_title("Capacity Breakdown",fontsize=13,fontweight="bold")
    ax.legend(); ax.axhline(y=750,color="k",ls="--",alpha=0.3); ax.grid(axis="y",alpha=0.3)
    plt.savefig(os.path.join(outdir,"capacity_breakdown.png"),dpi=150,bbox_inches="tight"); plt.close()

def plot_box(R, outdir):
    kpis=["jobs_completed","failures","service_level","total_cost","n_PM","availability_end"]
    titles=["Jobs","Failures","Service Level","Total Cost","PM Events","Availability"]
    names=list(R.keys()); fig,axes=plt.subplots(2,3,figsize=(18,10))
    for ax,k,t in zip(axes.flatten(),kpis,titles):
        data=[[r[k] for r in R[n]] for n in names]
        bp=ax.boxplot(data,patch_artist=True,tick_labels=[n.replace("_","\n") for n in names])
        for p,c in zip(bp["boxes"],COLORS[:len(names)]): p.set_facecolor(c); p.set_alpha(0.6)
        ax.set_title(t,fontsize=12,fontweight="bold"); ax.grid(axis="y",alpha=0.3)
    plt.tight_layout(); plt.savefig(os.path.join(outdir,"kpi_distributions.png"),dpi=150,bbox_inches="tight"); plt.close()

def main():
    pa=argparse.ArgumentParser(); pa.add_argument("--episodes",type=int,default=30)
    pa.add_argument("--outdir",default="outputs/baselines"); pa.add_argument("--config",default="configs/base.yaml")
    pa.add_argument("--stoch-level",type=int,default=1); args=pa.parse_args()
    with open(args.config) as f: config=yaml.safe_load(f)
    config["stochasticity_level"]=args.stoch_level
    outdir=args.outdir; os.makedirs(os.path.join(outdir,"per_baseline"),exist_ok=True)
    baselines=get_all_baselines(); baselines.append(MARLProxyBaseline())
    seeds=list(range(42,42+args.episodes)); env=ManufacturingEnv(config)
    print(f"\n{'='*70}\n  BASELINE BENCHMARK — {args.episodes} episodes x {len(baselines)} baselines\n{'='*70}\n")
    R={}
    for pol in baselines:
        pn=pol.name.replace(" ","_").replace("+","_").replace("(","").replace(")","").replace(",","")
        print(f"Running {pol.name}...")
        eps=[run_episode(env,pol,s) for s in seeds]; R[pn]=eps
        with open(os.path.join(outdir,"per_baseline",f"{pn}.csv"),"w",newline="") as f:
            w=csv.DictWriter(f,fieldnames=eps[0].keys()); w.writeheader(); w.writerows(eps)
    # Combined CSV
    with open(os.path.join(outdir,"per_episode_data.csv"),"w",newline="") as f:
        flds=["baseline"]+list(list(R.values())[0][0].keys())
        w=csv.DictWriter(f,fieldnames=flds); w.writeheader()
        for bn,eps in R.items():
            for ep in eps: row={"baseline":bn}; row.update(ep); w.writerow(row)
    # Summary
    kpi_keys=list(list(R.values())[0][0].keys()); kpi_keys.remove("seed")
    print(f"\n{'='*120}")
    hdr=f"{'KPI':<25}"; 
    for bn in R: hdr+=f" {bn:>20}"
    print(hdr); print("-"*120)
    for k in kpi_keys:
        line=f"{k:<25}"
        for bn in R: vs=[r[k] for r in R[bn]]; line+=f" {np.mean(vs):>12.1f}\u00b1{np.std(vs):>5.1f}"
        print(line)
    # Summary CSV
    with open(os.path.join(outdir,"summary_table.csv"),"w",newline="") as f:
        flds=["KPI"]+[f"{bn}_mean" for bn in R]+[f"{bn}_std" for bn in R]
        w=csv.DictWriter(f,fieldnames=flds); w.writeheader()
        for k in kpi_keys:
            row={"KPI":k}
            for bn in R: vs=[r[k] for r in R[bn]]; row[f"{bn}_mean"]=round(np.mean(vs),2); row[f"{bn}_std"]=round(np.std(vs),2)
            w.writerow(row)
    # Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Summary"
        hf=PatternFill(start_color="1a365d",end_color="1a365d",fill_type="solid"); hfont=Font(color="FFFFFF",bold=True,size=10)
        bns=list(R.keys()); hdrs=["KPI"]+[f"{b} mean" for b in bns]+[f"{b} std" for b in bns]
        for c,h in enumerate(hdrs,1): cell=ws.cell(row=1,column=c,value=h); cell.fill=hf; cell.font=hfont
        for ri,k in enumerate(kpi_keys,2):
            ws.cell(row=ri,column=1,value=k)
            for bi,bn in enumerate(bns):
                vs=[r[k] for r in R[bn]]; ws.cell(row=ri,column=2+bi,value=round(np.mean(vs),2))
                ws.cell(row=ri,column=2+len(bns)+bi,value=round(np.std(vs),2))
        ws2=wb.create_sheet("Raw_Data")
        flds2=["baseline"]+kpi_keys
        for c,h in enumerate(flds2,1): cell=ws2.cell(row=1,column=c,value=h); cell.fill=hf; cell.font=hfont
        ri=2
        for bn,eps in R.items():
            for ep in eps:
                ws2.cell(row=ri,column=1,value=bn)
                for ci,k in enumerate(kpi_keys,2): ws2.cell(row=ri,column=ci,value=round(ep[k],3))
                ri+=1
        wb.save(os.path.join(outdir,"summary_table.xlsx")); print(f"\nSaved Excel: {outdir}/summary_table.xlsx")
    except ImportError: print("  (openpyxl not installed)")
    # Plots
    print("\nGenerating plots..."); plot_bars(R,outdir); plot_radar(R,outdir); plot_capacity(R,outdir); plot_box(R,outdir)
    with open(os.path.join(outdir,"run_metadata.json"),"w") as f:
        json.dump({"timestamp":datetime.now().isoformat(),"episodes":args.episodes,"seeds":seeds,
            "stoch_level":args.stoch_level,"baselines":[p.name for p in baselines]},f,indent=2)
    print(f"\n{'='*70}\n  DONE — results in {outdir}/\n{'='*70}")

if __name__=="__main__": main()
