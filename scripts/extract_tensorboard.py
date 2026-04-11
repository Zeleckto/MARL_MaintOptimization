"""
scripts/extract_tensorboard.py
===============================
Extracts ALL TensorBoard scalar data from one or more run directories
into a single Excel/CSV file for analysis.

Usage:
    python scripts/extract_tensorboard.py --runs outputs/runs/phase1_ABC outputs/runs/phase1_XYZ
    python scripts/extract_tensorboard.py --rundir outputs/runs/   # all runs in folder
    python scripts/extract_tensorboard.py --rundir outputs/runs/ --outdir outputs/analysis

Output:
    outputs/analysis/
    ├── training_history.xlsx        # All runs, all metrics
    ├── training_history.csv         # Same as CSV
    ├── per_run/                     # Separate CSV per run
    │   ├── phase1_ABC.csv
    │   └── phase1_XYZ.csv
    ├── combined_episode_metrics.csv # Episode-level: jobs, fail, r1, r2, health per step
    └── training_curves.png          # Key metrics plotted over training
"""
import argparse, os, sys, glob
import csv
import numpy as np

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

try:
    from tensorboard.backend.event_processing.event_accumulator import EventAccumulator
except ImportError:
    print("Install tensorboard: pip install tensorboard")
    sys.exit(1)

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt


def extract_run(run_dir):
    """Extract all scalar data from a single TensorBoard run."""
    ea = EventAccumulator(run_dir)
    ea.Reload()

    tags = ea.Tags().get("scalars", [])
    if not tags:
        print(f"  WARNING: No scalar data in {run_dir}")
        return {}, []

    data = {}
    for tag in tags:
        events = ea.Scalars(tag)
        data[tag] = [(e.step, e.value, e.wall_time) for e in events]

    return data, tags


def build_episode_table(data, tags):
    """Build per-step table with all metrics aligned by step."""
    # Find all unique steps across all tags
    all_steps = set()
    for tag in tags:
        if tag in data:
            for step, val, wt in data[tag]:
                all_steps.add(step)

    all_steps = sorted(all_steps)
    if not all_steps:
        return []

    # Build rows
    rows = []
    for step in all_steps:
        row = {"step": step}
        for tag in tags:
            if tag in data:
                # Find value at this step
                for s, v, wt in data[tag]:
                    if s == step:
                        row[tag] = v
                        break
        rows.append(row)
    return rows


def plot_training_curves(all_rows, outdir, run_labels):
    """Plot key training metrics across all runs."""
    key_metrics = [
        ("episode/jobs_completed", "Jobs Completed"),
        ("episode/failures", "Failures"),
        ("episode/return_agent1", "Return Agent 1"),
        ("episode/return_agent2", "Return Agent 2"),
        ("train/entropy1", "Entropy Agent 1"),
        ("train/entropy2", "Entropy Agent 2"),
        ("train/actor1_loss", "Actor 1 Loss"),
        ("train/critic_loss", "Critic Loss"),
    ]

    # Filter to metrics that exist
    available = []
    for metric, label in key_metrics:
        for run_label, rows in all_rows.items():
            if any(metric in r for r in rows):
                available.append((metric, label))
                break

    if not available:
        print("  No plottable metrics found")
        return

    n = len(available)
    fig, axes = plt.subplots((n + 1) // 2, 2, figsize=(16, 4 * ((n + 1) // 2)))
    axes = axes.flatten()
    colors = plt.cm.tab10(np.linspace(0, 1, len(all_rows)))

    for ax, (metric, label) in zip(axes, available):
        for (run_label, rows), color in zip(all_rows.items(), colors):
            steps = [r["step"] for r in rows if metric in r]
            vals = [r[metric] for r in rows if metric in r]
            if steps:
                ax.plot(steps, vals, alpha=0.3, color=color, linewidth=0.5)
                # Rolling average
                if len(vals) > 20:
                    window = min(50, len(vals) // 5)
                    smoothed = np.convolve(vals, np.ones(window) / window, mode="valid")
                    ax.plot(steps[window - 1:], smoothed, color=color, linewidth=2,
                            label=run_label[:30])
                else:
                    ax.plot(steps, vals, color=color, linewidth=2, label=run_label[:30])
        ax.set_title(label, fontsize=11, fontweight="bold")
        ax.grid(alpha=0.3)
        ax.legend(fontsize=7)

    # Hide unused axes
    for i in range(len(available), len(axes)):
        axes[i].set_visible(False)

    plt.tight_layout()
    path = os.path.join(outdir, "training_curves.png")
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  Saved: {path}")


def main():
    pa = argparse.ArgumentParser()
    pa.add_argument("--runs", nargs="+", default=[], help="Specific run directories")
    pa.add_argument("--rundir", default=None, help="Parent dir containing run folders")
    pa.add_argument("--outdir", default="outputs/analysis")
    args = pa.parse_args()

    # Collect run directories
    run_dirs = list(args.runs)
    if args.rundir:
        for d in sorted(os.listdir(args.rundir)):
            full = os.path.join(args.rundir, d)
            if os.path.isdir(full):
                run_dirs.append(full)

    if not run_dirs:
        print("No runs found. Use --runs or --rundir.")
        return

    os.makedirs(args.outdir, exist_ok=True)
    os.makedirs(os.path.join(args.outdir, "per_run"), exist_ok=True)

    print(f"Found {len(run_dirs)} runs:")
    for d in run_dirs:
        print(f"  {d}")
    print()

    all_rows = {}
    all_tags = set()

    for run_dir in run_dirs:
        run_name = os.path.basename(run_dir)
        print(f"Extracting: {run_name}...")
        data, tags = extract_run(run_dir)
        if not tags:
            continue

        all_tags.update(tags)
        rows = build_episode_table(data, tags)
        all_rows[run_name] = rows

        # Save per-run CSV
        if rows:
            csv_path = os.path.join(args.outdir, "per_run", f"{run_name}.csv")
            fieldnames = ["step"] + sorted(set(k for r in rows for k in r if k != "step"))
            with open(csv_path, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
                writer.writeheader()
                writer.writerows(rows)
            print(f"  {len(rows)} data points, {len(tags)} metrics → {csv_path}")

    # Combined CSV with run label
    combined_path = os.path.join(args.outdir, "combined_all_runs.csv")
    all_fieldnames = ["run", "step"] + sorted(all_tags)
    with open(combined_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=all_fieldnames, extrasaction="ignore")
        writer.writeheader()
        for run_name, rows in all_rows.items():
            for row in rows:
                row["run"] = run_name
                writer.writerow(row)
    print(f"\nCombined CSV: {combined_path}")

    # Episode-level summary (jobs, failures, r1, r2)
    episode_metrics = ["episode/jobs_completed", "episode/failures",
                       "episode/return_agent1", "episode/return_agent2",
                       "episode/avg_health", "episode/n_PM",
                       "episode/service_level", "episode/maint_cost"]
    ep_path = os.path.join(args.outdir, "episode_summary.csv")
    with open(ep_path, "w", newline="") as f:
        fields = ["run", "step"] + [m.split("/")[-1] for m in episode_metrics]
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for run_name, rows in all_rows.items():
            for row in rows:
                has_ep = any(m in row for m in episode_metrics)
                if has_ep:
                    out = {"run": run_name, "step": row["step"]}
                    for m in episode_metrics:
                        key = m.split("/")[-1]
                        out[key] = row.get(m, "")
                    writer.writerow(out)
    print(f"Episode summary: {ep_path}")

    # Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()

        # Sheet 1: Episode summary per run
        for run_name, rows in all_rows.items():
            sheet_name = run_name[:31]  # Excel 31 char limit
            ws = wb.create_sheet(title=sheet_name)
            ep_rows = [r for r in rows if any(m in r for m in episode_metrics)]
            if not ep_rows:
                continue
            headers = ["step"] + [m.split("/")[-1] for m in episode_metrics if any(m in r for r in ep_rows)]
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font = Font(bold=True)
            for ri, row in enumerate(ep_rows, 2):
                ws.cell(row=ri, column=1, value=row["step"])
                for ci, m in enumerate(episode_metrics, 2):
                    if m in row:
                        ws.cell(row=ri, column=ci, value=round(row[m], 3))

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        xlsx_path = os.path.join(args.outdir, "training_history.xlsx")
        wb.save(xlsx_path)
        print(f"Excel: {xlsx_path}")
    except ImportError:
        print("  (openpyxl not installed — skipping Excel)")

    # Plot
    print("\nGenerating training curves...")
    plot_training_curves(all_rows, args.outdir, list(all_rows.keys()))

    print(f"\n{'='*60}")
    print(f"  All outputs in: {args.outdir}/")
    print(f"  Runs processed: {len(all_rows)}")
    print(f"  Total metrics: {len(all_tags)}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
